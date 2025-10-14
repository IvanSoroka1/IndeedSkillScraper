from flask import Flask, request
from collections import defaultdict, Counter
from openpyxl import Workbook
from openpyxl.styles import Font

app = Flask(__name__)


from openpyxl import Workbook
from openpyxl.styles import Font

def create_jobs_sheet(wb, jobs):
    ws = wb.create_sheet(title="Jobs")
    ws.append(["Title", "Company", "Skills", "Location", "Remote", "Id"])
    
    # Make header bold
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for job in jobs:
        # Append all fields except hyperlink first
        ws.append([
                   job.get("Title", ""),
                   job.get("Company", ""),
                   job.get("Skills", ""),
                   job.get("Location", ""),
                   job.get("Remote", ""),
                   job.get("Id", ""),
                   ])

        row_idx = ws.max_row
        url = f"https://ca.indeed.com/viewjob?jk={job.get('Id','')}"
        cell = ws.cell(row=row_idx, column=1)
        cell.hyperlink = url
        cell.style = "Hyperlink"


@app.route('/isOnline', methods=['GET'])
def isOnline():
    return {"status": "success"}, 200

@app.route('/upload_jobs', methods=['POST'])
def upload_jobs():
    data = request.get_json()
    jobs = data.get("jobs", [])

    # Build skill → list of jobs mapping
    skill_to_jobs = {}
    for job in jobs:
        job_id = job.get("Id", "")
        job_title = job.get("Title", "Unknown Title")
        job_company = job.get("Company", "Unknown Company")
        skills = job.get("Skills", "")
        if not skills:
            continue
        for skill in [s.strip() for s in skills.split(";") if s.strip()]:
            if skill not in skill_to_jobs:
                skill_to_jobs[skill] = []
            skill_to_jobs[skill].append((job_title, job_company, job_id))

    # Create Excel workbook
    wb = Workbook()

    # Create each worksheet
    create_jobs_sheet(wb, jobs)

    # Save workbook
    output_path = "skill_counts.xlsx"
    wb.save(output_path)

    return {"status": "success", "message": f"Saved to {output_path}"}, 200

if __name__ == '__main__':
    app.run(port=5000)
