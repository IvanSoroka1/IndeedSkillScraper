from collections import Counter, defaultdict
from itertools import combinations
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

def create_skill_counts_sheet(wb, skill_to_jobs):
    """Creates the first worksheet: Skill counts with job hyperlinks"""
    ws = wb.create_sheet(title="Skill Counts")
    ws.title = "Skill Counts"

    # Header
    ws.append(["Skill", "Count"])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Sort skills by count (descending)
    sorted_skills = sorted(skill_to_jobs.items(), key=lambda x: len(x[1]), reverse=True)

    # Fill rows
    for skill, jobs_list in sorted_skills:
        row = [skill, len(jobs_list)]
        ws.append(row)
        row_idx = ws.max_row
        col_idx = 3  # start placing jobs from column C
        for title, company, job_id in jobs_list:
            cell = ws.cell(row=row_idx, column=col_idx)
            url = f"https://ca.indeed.com/viewjob?jk={job_id}"
            cell.value = f"{company} - {title}"
            cell.hyperlink = url
            cell.style = "Hyperlink"
            col_idx += 1


def create_skill_combinations_sheet(wb, jobs, n=2):
    """
    Creates a worksheet listing the most common n-skill combinations across jobs,
    ignoring combinations that appear only once (for speed and clarity).

    Args:
        wb: openpyxl Workbook object
        jobs: list of job dictionaries, each with a 'Skills' field (semicolon-separated)
        n: size of combinations (e.g., 2 for pairs, 3 for triplets, etc.)
    
    Returns:
        True if any combinations were added, False otherwise
    """
    sheet_name = f"Top {n}-Skill Combos"
    ws = wb.create_sheet(title=sheet_name)

    # Header row
    header = [f"Skill {i+1}" for i in range(n)] + ["Co-occurrence Count"]
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Count co-occurrences of n-skill combinations
    combo_counter = Counter()
    for job in jobs:
        skills = [s.strip() for s in (job.get("Skills") or "").split(";") if s.strip()]
        if len(skills) >= n:
            for combo in combinations(sorted(set(skills)), n):
                combo_counter[combo] += 1

    # Keep only combinations that appear more than once
    frequent_combos = [(combo, count) for combo, count in combo_counter.items() if count > 1]
    
    if not frequent_combos:
        return False

    # Sort by frequency (descending)
    for combo, count in sorted(frequent_combos, key=lambda x: x[1], reverse=True):
        ws.append(list(combo) + [count])

    return True



def create_co_occurrences_sheet(wb, jobs):
    """Creates the second worksheet: Skill co-occurrences horizontally"""
    ws = wb.create_sheet(title="Skill Co-Occurrences")

    # Parse jobs into sets of skills
    job_skills = [set(s.strip() for s in job.get("Skills", "").split(";") if s.strip()) for job in jobs]

    # Build co-occurrence Counter
    co_occurrence = defaultdict(Counter)
    for skills in job_skills:
        for skill in skills:
            for other_skill in skills:
                if skill != other_skill:
                    co_occurrence[skill][other_skill] += 1

    # Header row
    ws.append(["Skill"] + ["Co-occurring Skill (Count)"] * 10)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Fill rows
    for skill, counter in co_occurrence.items():
        row = [skill]
        for other_skill, count in counter.most_common(10):  # top 10
            row.append(f"{other_skill} ({count})")
        ws.append(row)

from openpyxl import load_workbook

def upload_jobs_from_excel(input_path):
    # Load the existing Excel file
    wb = load_workbook(input_path)
    ws = wb["Jobs"]

    # Read column headers
    headers = [cell.value for cell in ws[1]]

    # Convert each row to a dictionary
    jobs = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        job = dict(zip(headers, row))
        jobs.append(job)

    # Build skill → list of jobs mapping
    skill_to_jobs = {}
    for job in jobs:
        job_id = job.get("Id", "")
        job_title = job.get("Title", "Unknown Title")
        job_company = job.get("Company", "Unknown Company")
        skills = job.get("Skills", "")
        if not skills:
            continue
        for skill in [s.strip() for s in skills.split(";") if s.strip()]:
            if skill not in skill_to_jobs:
                skill_to_jobs[skill] = []
            skill_to_jobs[skill].append((job_title, job_company, job_id))

    # Use the same workbook to add sheets
    create_skill_counts_sheet(wb, skill_to_jobs)

    # keepGoing = True
    # i = 2
    # while keepGoing:
    #     keepGoing = create_skill_combinations_sheet(wb, jobs, i)
    #     print(f"✅ Created sheet: Top {i}-Skill Combos")
    #     i+=1
    for i in range(2, 8):
        create_skill_combinations_sheet(wb, jobs, i)
    #create_co_occurrences_sheet(wb, jobs)

    # Save back to the same file (or a new one if preferred)
    wb.save(input_path)
    print(f"✅ Updated {input_path}")


if __name__ == "__main__":
    upload_jobs_from_excel("skill_counts.xlsx")